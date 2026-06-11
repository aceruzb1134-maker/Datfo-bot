"""
Excel parser for FOM Group DATFO bot.

Svod sheet structure (row 3 = headers, row 5+ = data):
  Col B (2)  = ИНН
  Col C (3)  = Официальное юридическое название
  Col D (4)  = Аптеки (название аптеки)
  Col K (11) = ПЛАН-I-Q (общий план квартала)
  Col L (12) = План Январь
  Col P (16) = План Февраль
  Col T (20) = План Март
  Col O (15) = Бонус Январь
  Col S (19) = Бонус Февраль  
  Col W (23) = Бонус Март
  Col Z (26) = Total Bonus

Yanvar/Fevral/Mart sheet structure (row 1 = headers, row 2+ = data):
  Col E (5)  = ИНН
  Col C (3)  = Лекарство (название препарата)
  Col L (12) = Кол-во проданных товаров (план/факт)
  Col N (14) = Сип (цена)
"""

import openpyxl
from io import BytesIO


def parse_svod(file_bytes: bytes) -> dict | None:
    """
    Parse Svod sheet — returns per-pharmacy quarterly plan.
    Returns:
    {
        "producer": "Medexport",
        "period": "I квартал 2026",
        "pharmacies": {
            "302341262": {
                "inn": "302341262",
                "legal_name": "TURDI RASULOVICH XK",
                "name": "TURDI RASULOVICH XK",
                "plan_total": 8250000,
                "plan_jan": 2750000,
                "plan_feb": 2750000,
                "plan_mar": 2750000,
                "bonus_total": 0,
            },
            ...
        }
    }
    """
    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)

        if "Svod" not in wb.sheetnames:
            return None

        ws = wb["Svod"]

        # Row 1 col A has period name like "datfo-I-Q-2026"
        period_raw = ws.cell(row=1, column=1).value or ""
        period = str(period_raw).replace("datfo-", "").replace("-", " ").strip()

        # Row 2 col J has producer name
        producer = str(ws.cell(row=2, column=10).value or "").strip()

        pharmacies = {}
        for row in ws.iter_rows(min_row=5, values_only=True):
            inn = row[1]   # col B
            if inn is None or str(inn).strip() in ("-", "", "None"):
                continue
            try:
                inn_str = str(int(float(str(inn))))
            except (ValueError, TypeError):
                inn_str = str(inn).strip()

            legal_name = str(row[2] or "").strip()
            name       = str(row[3] or "").strip() or legal_name

            def safe_float(val):
                try:
                    return float(val) if val not in (None, "", "-") else 0
                except (ValueError, TypeError):
                    return 0

            plan_total = safe_float(row[10])  # col K
            plan_jan   = safe_float(row[11])  # col L
            plan_feb   = safe_float(row[15])  # col P
            plan_mar   = safe_float(row[19])  # col T
            bonus_jan  = safe_float(row[14])  # col O
            bonus_feb  = safe_float(row[18])  # col S
            bonus_mar  = safe_float(row[22])  # col W
            bonus_total= safe_float(row[25])  # col Z

            pharmacies[inn_str] = {
                "inn":        inn_str,
                "legal_name": legal_name,
                "name":       name,
                "plan_total": plan_total,
                "plan_jan":   plan_jan,
                "plan_feb":   plan_feb,
                "plan_mar":   plan_mar,
                "bonus_jan":  bonus_jan,
                "bonus_feb":  bonus_feb,
                "bonus_mar":  bonus_mar,
                "bonus_total":bonus_total,
            }

        if not pharmacies:
            return None

        return {
            "producer":   producer or "DATFO",
            "period":     period or "I квартал",
            "pharmacies": pharmacies,
        }

    except Exception as e:
        print(f"parse_svod error: {e}")
        return None


def parse_month_sheet(file_bytes: bytes, sheet_name: str) -> dict:
    """
    Parse Yanvar/Fevral/Mart sheet.
    Returns: { inn_str: [ {sku, qty, price}, ... ], ... }
    """
    result = {}
    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            return result

        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            inn = row[4]   # col E
            sku = row[2]   # col C — Лекарство
            qty = row[11]  # col L — Кол-во
            price = row[13] # col N — Сип

            if inn is None or sku is None:
                continue
            try:
                inn_str = str(int(float(str(inn))))
            except (ValueError, TypeError):
                continue

            try:
                qty_int = int(float(str(qty))) if qty not in (None, "", "-") else 0
            except (ValueError, TypeError):
                qty_int = 0

            try:
                price_f = float(price) if price not in (None, "", "-") else 0
            except (ValueError, TypeError):
                price_f = 0

            if qty_int <= 0:
                continue

            if inn_str not in result:
                result[inn_str] = []
            result[inn_str].append({
                "sku":   str(sku).strip(),
                "qty":   qty_int,
                "price": price_f,
            })

    except Exception as e:
        print(f"parse_month_sheet({sheet_name}) error: {e}")

    return result


def fmt_money(amount: float) -> str:
    """Format number as UZS: 2 750 000 сум"""
    if not amount:
        return "—"
    return f"{int(amount):,}".replace(",", " ") + " сум"


def format_pharmacy_plan_message(pharmacy: dict, items_jan: list, items_feb: list, items_mar: list) -> str:
    """
    Full message sent to pharmacy with quarterly plan.
    """
    lines = [
        f"📋 *Ваш план продаж*\n",
        f"📅 Квартал: план на 3 месяца",
        f"💰 Общий план: *{fmt_money(pharmacy['plan_total'])}*",
    ]

    if pharmacy.get("bonus_total"):
        lines.append(f"🎁 Бонус при выполнении: *{fmt_money(pharmacy['bonus_total'])}*")

    lines.append("")

    # Monthly breakdown
    months = [
        ("Январь", pharmacy["plan_jan"], pharmacy.get("bonus_jan", 0), items_jan),
        ("Февраль", pharmacy["plan_feb"], pharmacy.get("bonus_feb", 0), items_feb),
        ("Март", pharmacy["plan_mar"], pharmacy.get("bonus_mar", 0), items_mar),
    ]

    for month_name, plan, bonus, items in months:
        lines.append(f"📅 *{month_name}* — {fmt_money(plan)}")
        if bonus:
            lines.append(f"   🎁 Бонус: {fmt_money(bonus)}")
        if items:
            for item in items[:10]:  # max 10 препаратов
                lines.append(f"   • {item['sku']} — *{item['qty']} уп.*")
            if len(items) > 10:
                lines.append(f"   _...и ещё {len(items)-10} препаратов_")
        lines.append("")

    lines.append("❓ *Сможете выполнить этот план?*")
    return "\n".join(lines)


def format_my_plan_message(pharmacy: dict, items_jan: list, items_feb: list, items_mar: list) -> str:
    """Message for /myplan command — same as above but without question."""
    lines = [
        f"📋 *Ваш план продаж*\n",
        f"💰 Общий план квартала: *{fmt_money(pharmacy['plan_total'])}*",
    ]
    if pharmacy.get("bonus_total"):
        lines.append(f"🎁 Бонус при выполнении: *{fmt_money(pharmacy['bonus_total'])}*")
    lines.append("")

    months = [
        ("Январь", pharmacy["plan_jan"], pharmacy.get("bonus_jan", 0), items_jan),
        ("Февраль", pharmacy["plan_feb"], pharmacy.get("bonus_feb", 0), items_feb),
        ("Март", pharmacy["plan_mar"], pharmacy.get("bonus_mar", 0), items_mar),
    ]
    for month_name, plan, bonus, items in months:
        lines.append(f"📅 *{month_name}* — {fmt_money(plan)}")
        if bonus:
            lines.append(f"   🎁 Бонус: {fmt_money(bonus)}")
        if items:
            for item in items[:10]:
                lines.append(f"   • {item['sku']} — *{item['qty']} уп.*")
            if len(items) > 10:
                lines.append(f"   _...и ещё {len(items)-10} препаратов_")
        lines.append("")

    return "\n".join(lines)


def format_admin_preview(data: dict) -> str:
    """Preview for admin before broadcasting."""
    ph_list = list(data["pharmacies"].values())
    first = ph_list[0]
    lines = [
        f"👀 *Предпросмотр кампании*\n",
        f"🏭 Производитель: {data['producer']}",
        f"📅 Период: {data['period']}",
        f"🏪 Аптек в файле: {len(ph_list)}\n",
        f"*Пример — первая аптека:*",
        f"🔑 ИНН: {first['inn']}",
        f"🏪 {first['name']}",
        f"💰 План квартала: {fmt_money(first['plan_total'])}",
        f"   Январь: {fmt_money(first['plan_jan'])}",
        f"   Февраль: {fmt_money(first['plan_feb'])}",
        f"   Март: {fmt_money(first['plan_mar'])}",
        f"\n✅ Всё верно? Запустить рассылку?",
    ]
    return "\n".join(lines)
